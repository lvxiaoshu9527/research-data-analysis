import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
import tkinter as tk
from tkinter import filedialog, simpledialog
import sys
from scipy.optimize import curve_fit

def select_folder():
    """打开文件夹选择对话框"""
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="选择数据文件夹")
    if not folder_path:
        print("没有选择文件夹，程序退出。")

        sys.exit()
    return folder_path

def lorentzian(x, amp, cen, wid):
    """定义洛伦兹函数"""
    return (amp * wid**2 / ((x - cen)**2 + wid**2))

def get_wavelength_range(filename):
    """根据文件名确定波长范围"""
    ranges = {
        "(6,4)": (871.2951, 961.1901),
        "(6,5)-S7": (895, 1100.975),
        "(6,5)": (950, 1100),
        "(7,3)": (961.1901, 1074),
        "(7,5)": (980.129, 1150.973),
        "(8,3)": (926.6856, 1050.47),
        "(9,1)": (890.3606, 1021.353)
    }

    for key in ranges:
        if key in filename:
            return ranges[key]
    return None

def get_output_filename():
    """获取用户输入的输出文件名"""
    root = tk.Tk()
    root.withdraw()
    filename = simpledialog.askstring("文件名", "请输入输出Excel文件名 (不需要.xlsx后缀)：",
                                      initialvalue="output_data")
    if not filename:
        filename = "output_data"
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    return filename

def aggregate_fluorescence_data(directory_path, output_file):
    """汇总指定目录下的荧光数据，并保存到 Excel 文件中。"""
    print("正在汇总荧光数据...")
    data_frames = []
    for filename in os.listdir(directory_path):
        if filename.endswith("660.txt"):
            file_path = os.path.join(directory_path, filename)
            df = pd.read_csv(file_path, sep='\t', header=None, names=["Wavelength", "Intensity"])
            df = df[(df['Wavelength'] >= 900) & (df['Wavelength'] <= 1400)]
            df = df.set_index('Wavelength').rename(columns={"Intensity": os.path.splitext(filename)[0]})
            data_frames.append(df)

    if not data_frames:
        print("警告：未找到符合条件的数据文件！")
        sys.exit()

    all_data = pd.concat(data_frames, axis=1)
    all_data.index.name = "Wavelength"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        all_data.to_excel(writer, sheet_name='All')
    print("荧光数据汇总完成。")

def extract_max_values(file_path):
    """从 Excel 文件中提取最大值并保存到新的 sheet 中。"""
    print("正在提取峰值数据...")
    df = pd.read_excel(file_path, sheet_name='All', index_col='Wavelength')
    results = []
    for col in df.columns:
        max_value = df[col].max()
        min_value = df[col].min()
        peak_position = df[df[col] == max_value].index[0]
        results.append({'Name': col, 'Peak Position': peak_position, 'Intensity': max_value - min_value})
    new_df = pd.DataFrame(results).dropna(how='all')
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        new_df.to_excel(writer, sheet_name='Max_Value', index=False)
    print("峰值数据提取完成。")

def calculate_intensity_ratios(file_path):
    """计算相对于参考样本(0__660)的(I-I0)/I0比值。"""
    print("正在计算强度比值...")
    wb = load_workbook(file_path)
    sheet = wb['Max_Value']
    ref_name = "0__660"
    ref_row = None
    for row in sheet.iter_rows(min_row=2):
        if row[0].value and row[0].value.startswith(ref_name):
            ref_row = row
            break

    if ref_row is None:
        print(f"警告：未找到参考样本 '{ref_name}'！")
        return

    data_list = []
    for row in sheet.iter_rows(min_row=2):
        name = row[0].value
        intensity = row[2].value
        try:
            intensity = float(intensity)
        except (ValueError, TypeError):
            print(f"警告：样本 '{name}' 的强度值 '{intensity}' 无法转换为数字，已跳过。")
            continue

        data_list.append({'Name': name, 'Intensity': intensity})

    df = pd.DataFrame(data_list)
    ref_samples = df[df['Name'].str.startswith(ref_name, na=False)]

    if ref_samples.empty:
          print(f"警告: 无法在数据中找到参考样本{ref_name}!")
          return

    i0 = ref_samples['Intensity'].mean()
    ratios = (df['Intensity'] - i0) / i0
    sheet['D1'] = '(I-I0)/I0'
    for i, ratio in enumerate(ratios, start=2):
        sheet[f'D{i}'] = ratio

    wb.save(file_path)
    print("强度比值计算完成。")

def format_fonts(sheet):
    """格式化Excel表格样式"""
    header_font = Font(name='Arial', bold=True)
    body_font = Font(name='Times New Roman')
    right_alignment = Alignment(horizontal='right')
    no_border = Border(left=Side(style='none'),
                       right=Side(style='none'),
                       top=Side(style='none'),
                       bottom=Side(style='none'))

    for row in sheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = header_font
            else:
                cell.font = body_font
            cell.alignment = right_alignment
            cell.border = no_border

def process_sheet_data(df, wave_range, sheet_name):
    """处理单个 sheet 的数据拟合"""
    mask = (df.index >= wave_range[0]) & (df.index <= wave_range[1])
    filtered_df = df[mask]

    x = filtered_df.index.values
    results = pd.DataFrame(columns=["Column", "Amplitude", "Center", "Width"])
    fit_data = pd.DataFrame({'x': x})

    for column in filtered_df.columns:
        y = filtered_df[column].values

        # --- 初始参数猜测 (改进) ---
        max_index = np.argmax(y)
        max_y = y[max_index]
        max_x = x[max_index]

        half_max = max_y / 2.0
        left_side = y[:max_index]
        right_side = y[max_index:]
        left_diff = np.abs(left_side - half_max)
        right_diff = np.abs(right_side - half_max)
        left_idx = np.argmin(left_diff)
        right_idx = np.argmin(right_diff)
        left_x = x[left_idx]
        right_x = x[max_index + right_idx]
        fwhm = right_x - left_x
        wid_guess = fwhm / 2.0

        initial_guess = [max_y, max_x, wid_guess]
        # --- 初始参数猜测结束 ---

        try:
            # --- 原代码结构 ---
            popt, _ = curve_fit(lorentzian, x, y, p0=initial_guess)
            results = pd.concat([results, pd.DataFrame({
                "Column": [column],
                "Amplitude": [popt[0]],
                "Center": [popt[1]],
                "Width": [popt[2]]
            })], ignore_index=True)

            y_fit = lorentzian(x, *popt)
            fit_data[column] = y_fit
            # --- 原代码结构结束 ---

        except Exception as e:
            print(f"拟合 {column} 时出错: {str(e)}")
            continue

    return results, fit_data

def process_and_fit_data(file_path):
    """处理数据并进行洛伦兹拟合 (只针对 Lorentz)"""
    print("正在进行数据提取和洛伦兹拟合...")
    excel_filename = os.path.basename(file_path)
    df = pd.read_excel(file_path, sheet_name='All', index_col='Wavelength')
    lorentz_wave_range = get_wavelength_range(excel_filename)
    if lorentz_wave_range is None:
        print(f"警告：在文件名 '{excel_filename}' 中未找到匹配的波长范围！")
        return

    lorentz_results, lorentz_fit_data = process_sheet_data(df, lorentz_wave_range, 'Lorentz')

    if not lorentz_results.empty:
        reference_center = lorentz_results['Center'].iloc[0]
        lorentz_results['Δλ'] = lorentz_results['Center'] - reference_center
    else:
        print("警告：未成功进行任何拟合，无法计算 Δλ！")
        return

    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            lorentz_results.to_excel(writer, sheet_name="Lorentz", index=False)
            lorentz_fit_data.to_excel(writer, sheet_name="Fitted Data", index=False)
            wb = writer.book
            for sheet_name in ['Lorentz', 'Fitted Data']:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for cell in sheet[1]:
                        cell.font = Font(name='Arial')
                    for row in range(2, sheet.max_row + 1):
                        for col in range(1, sheet.max_column + 1):
                            sheet.cell(row=row, column=col).font = Font(name='Times New Roman')

        print(f"拟合结果已保存到：{file_path} 的 'Lorentz' 和 'Fitted Data' sheet 中")
        print(f"参考中心波长为：{reference_center:.2f}")

    except Exception as e:
        print(f"保存结果时出错: {str(e)}")

def modify_excel_data(file_path, has_conc):
    """根据是否有conc修改Excel数据, 针对特定sheet和行列"""
    print("正在修改Excel数据...")
    wb = load_workbook(file_path)
    print("所有 Sheet 名称:", wb.sheetnames)

    if has_conc:
        replacements = {
            '0__660': '0', '1__660': '10', '2__660': '30',
            '3__660': '50', '4__660': '70', '5__660': '90',
            '6__660': '110'
        }
    else:
        replacements = {
        '0__660': '0', '1__660': '20', '2__660': '40',
        '3__660': '60', '4__660': '80', '5__660': '100',
        '6__660': '120'
    }

    sheet_names = wb.sheetnames

    if 'All' in sheet_names:
        sheet1 = wb['All']
        print("正在处理 Sheet1 (All) 的第一行...")
        for cell in sheet1[1]:
            print(f"  单元格: {cell.coordinate}, 值: {cell.value}, 类型: {type(cell.value)}")
            if cell.value is not None and isinstance(cell.value, str):
                original_value = cell.value
                for old_val, new_val in replacements.items():
                    if cell.value == old_val:
                        cell.value = new_val
                        cell.number_format = numbers.FORMAT_TEXT
                print(f"    单元格 {cell.coordinate} 已修改: {original_value} -> {cell.value}")

    if 'Max_Value' in sheet_names:
        sheet2 = wb['Max_Value']
        print("正在处理 Sheet2 (Max_Value) 的第一列...")
        for row in sheet2.iter_rows():
            cell = row[0]
            print(f"  单元格: {cell.coordinate}, 值: {cell.value}, 类型: {type(cell.value)}")
            if cell.value is not None and isinstance(cell.value, str):
                original_value = cell.value
                for old_val, new_val in replacements.items():
                    if cell.value == old_val:
                        cell.value = new_val
                        cell.number_format = numbers.FORMAT_TEXT
                print(f"    单元格 {cell.coordinate} 已修改: {original_value} -> {cell.value}")

    if 'Lorentz' in sheet_names:
        sheet3 = wb['Lorentz']
        print("正在处理 Sheet3 (Lorentz) 的第一列...")
        for row in sheet3.iter_rows():
            cell = row[0]
            print(f"  单元格: {cell.coordinate}, 值: {cell.value}, 类型: {type(cell.value)}")
            if cell.value is not None and isinstance(cell.value, str):
                original_value = cell.value
                for old_val, new_val in replacements.items():
                    if cell.value == old_val:
                        cell.value = new_val
                        cell.number_format = numbers.FORMAT_TEXT
                print(f"    单元格 {cell.coordinate} 已修改: {original_value} -> {cell.value}")

    if 'Fitted Data' in sheet_names:
        sheet4 = wb['Fitted Data']
        print("正在处理 Sheet4 (Fitted Data) 的第一行...")
        for cell in sheet4[1]:
            print(f"  单元格: {cell.coordinate}, 值: {cell.value}, 类型: {type(cell.value)}")
            if cell.value is not None and isinstance(cell.value, str):
                original_value = cell.value
                for old_val, new_val in replacements.items():
                    if cell.value == old_val:
                        cell.value = new_val
                        cell.number_format = numbers.FORMAT_TEXT
                print(f"    单元格 {cell.coordinate} 已修改: {original_value} -> {cell.value}")

    wb.save(file_path)
    print("Excel数据修改完成。")

def main():
    print("欢迎使用荧光数据分析程序！")
    print("\n请在弹出的对话框中选择数据文件夹...")

    folder_path = select_folder()
    print(f"\n已选择文件夹: {folder_path}")

    output_filename = get_output_filename()
    output_file = os.path.join(folder_path, output_filename)
    print(f"输出文件将保存为: {output_file}")

    try:
        aggregate_fluorescence_data(folder_path, output_file)
        extract_max_values(output_file)
        calculate_intensity_ratios(output_file)
        process_and_fit_data(output_file)
        has_conc = "conc" in folder_path.lower()
        modify_excel_data(output_file, has_conc)

        print("正在格式化Excel...")
        wb = load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            format_fonts(wb[sheet_name])
        wb.save(output_file)

        print(f"\n处理完成！结果已保存到：{output_file}")
        print("\n按回车键退出程序...")
        input()

    except Exception as e:
        print(f"\n处理过程中出现错误：{str(e)}")
        print("\n按回车键退出程序...")
        input()

if __name__ == "__main__":
    main()

