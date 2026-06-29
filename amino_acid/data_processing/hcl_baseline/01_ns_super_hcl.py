import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
import tkinter as tk
from tkinter import filedialog, simpledialog
import sys
from scipy.optimize import curve_fit
import re


def select_folder():
    """打开文件夹选择对话框"""
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="选择数据文件夹")
    if not folder_path:
        print("没有选择文件夹，程序退出。")
        sys.exit()
    return folder_path


def lorentzian(x, amp, cen, wid):
    """定义洛伦兹函数"""
    return (amp * wid ** 2 / ((x - cen) ** 2 + wid ** 2))


def get_wavelength_range(filename):
    """根据文件名确定波长范围"""
    # 注意：请确保这些键值对能匹配你的文件夹名字
    ranges = {
        "(6,4)": (871.2951, 961.1901),
        "(6,5)": (950, 1100),
        "(7,3)": (961.1901, 1074),
        "(7,5)": (980.129, 1150.973),
        "(8,3)": (926.6856, 1050.47),
        "(9,1)": (890.3606, 1021.353)
    }
    for key in ranges:
        if key in filename:
            return ranges[key]
    # 默认范围
    return (900, 1400)


def get_output_filename():
    """获取用户输入的输出文件名"""
    root = tk.Tk()
    root.withdraw()
    filename = simpledialog.askstring("文件名", "请输入输出Excel文件名 (不需要.xlsx后缀)：",
                                      initialvalue="output_data")
    if not filename:
        filename = "output_data"
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    return filename


def natural_sort_key(s):
    """自然排序键，确保 10 在 2 后面"""
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', s)]


def aggregate_fluorescence_data(directory_path, output_file):
    """汇总指定目录下的荧光数据"""
    print("正在汇总荧光数据 (优先读取 660.txt)...")

    all_files = os.listdir(directory_path)

    # === 修正点：只找 660.txt 文件，忽略 NIR ===
    target_files = [f for f in all_files if f.endswith("660.txt")]

    # 按数字自然排序 (0, 1, 2, ... 10, 11)
    target_files.sort(key=natural_sort_key)

    if not target_files:
        print("错误：文件夹中未找到 *660.txt 文件！")
        sys.exit()

    print(f"找到 {len(target_files)} 个数据文件 (格式: 660.txt)，开始处理...")

    data_frames = []
    for filename in target_files:
        file_path = os.path.join(directory_path, filename)
        try:
            # 智能读取 (自动识别制表符或空格)
            df = pd.read_csv(file_path, sep=None, engine='python', header=None)

            # 确保取前两列 (波长, 强度)
            if df.shape[1] < 2:
                print(f"跳过文件 {filename}: 列数不足")
                continue

            df = df.iloc[:, :2]
            df.columns = ["Wavelength", "Intensity"]

            # 截取波长范围 900-1400 nm (根据你的旧代码习惯)
            df = df[(df['Wavelength'] >= 900) & (df['Wavelength'] <= 1400)]

            # 使用文件名作为列名 (去掉后缀)
            col_name = os.path.splitext(filename)[0]
            df = df.set_index('Wavelength').rename(columns={"Intensity": col_name})

            # 去除重复索引
            df = df[~df.index.duplicated(keep='first')]

            data_frames.append(df)
        except Exception as e:
            print(f"读取文件 {filename} 失败: {e}")

    if not data_frames:
        print("错误：没有成功读取任何数据！")
        sys.exit()

    all_data = pd.concat(data_frames, axis=1)
    all_data.index.name = "Wavelength"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        all_data.to_excel(writer, sheet_name='All')
    print("荧光数据汇总完成。")


def extract_max_values(file_path):
    """从 Excel 文件中提取最大值"""
    print("正在提取峰值数据...")
    df = pd.read_excel(file_path, sheet_name='All', index_col='Wavelength')
    results = []
    for col in df.columns:
        max_value = df[col].max()
        min_value = df[col].min()
        peak_position = df[col].idxmax()
        results.append({'Name': col, 'Peak Position': peak_position, 'Intensity': max_value - min_value})
    new_df = pd.DataFrame(results).dropna(how='all')
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        new_df.to_excel(writer, sheet_name='Max_Value', index=False)
    print("峰值数据提取完成。")


def calculate_intensity_ratios(file_path):
    """计算相对于第一个样本(通常是0)的(I-I0)/I0比值"""
    print("正在计算强度比值...")
    wb = load_workbook(file_path)
    sheet = wb['Max_Value']

    # 默认取第二行（即排序后的第一个数据，通常是0浓度）
    first_data_row = sheet[2]
    ref_name = first_data_row[0].value

    try:
        i0 = float(first_data_row[2].value)  # Intensity 列
        print(f"基准样本锁定为: {ref_name}, I0 = {i0}")
    except (ValueError, TypeError):
        print("错误：无法读取基准样本的强度值！")
        return

    sheet['D1'] = '(I-I0)/I0'

    for row in sheet.iter_rows(min_row=2):
        intensity_cell = row[2]  # Intensity
        output_cell = row[3]  # D列

        try:
            current_i = float(intensity_cell.value)
            ratio = (current_i - i0) / i0
            output_cell.value = ratio
        except (ValueError, TypeError):
            continue

    wb.save(file_path)
    print("强度比值计算完成。")


def format_fonts(sheet):
    """格式化Excel表格样式"""
    header_font = Font(name='Arial', bold=True)
    body_font = Font(name='Times New Roman')
    right_alignment = Alignment(horizontal='right')
    no_border = Border(left=Side(style='none'), right=Side(style='none'), top=Side(style='none'),
                       bottom=Side(style='none'))

    for row in sheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = header_font
            else:
                cell.font = body_font
            cell.alignment = right_alignment
            cell.border = no_border


def process_sheet_data(df, wave_range, sheet_name):
    """处理单个 sheet 的数据拟合"""
    mask = (df.index >= wave_range[0]) & (df.index <= wave_range[1])
    filtered_df = df[mask]

    x = filtered_df.index.values
    results = pd.DataFrame(columns=["Column", "Amplitude", "Center", "Width"])
    fit_data = pd.DataFrame({'x': x})

    for column in filtered_df.columns:
        y = filtered_df[column].values
        max_index = np.argmax(y)
        max_y = y[max_index]
        max_x = x[max_index]
        initial_guess = [max_y, max_x, 10.0]

        try:
            popt, _ = curve_fit(lorentzian, x, y, p0=initial_guess, maxfev=5000)
            popt[2] = abs(popt[2])

            results = pd.concat([results, pd.DataFrame({
                "Column": [column],
                "Amplitude": [popt[0]],
                "Center": [popt[1]],
                "Width": [popt[2]]
            })], ignore_index=True)

            y_fit = lorentzian(x, *popt)
            fit_data[column] = y_fit

        except Exception as e:
            print(f"拟合 {column} 时出错: {str(e)}")
            continue

    return results, fit_data


def process_and_fit_data(file_path):
    """处理数据并进行洛伦兹拟合"""
    print("正在进行数据提取和洛伦兹拟合...")
    excel_filename = os.path.basename(file_path)
    df = pd.read_excel(file_path, sheet_name='All', index_col='Wavelength')

    lorentz_wave_range = get_wavelength_range(excel_filename)
    print(f"使用的波长拟合范围: {lorentz_wave_range}")

    lorentz_results, lorentz_fit_data = process_sheet_data(df, lorentz_wave_range, 'Lorentz')

    if not lorentz_results.empty:
        reference_center = lorentz_results['Center'].iloc[0]
        lorentz_results['Δλ'] = lorentz_results['Center'] - reference_center
    else:
        print("警告：未成功进行任何拟合！")
        return

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        lorentz_results.to_excel(writer, sheet_name="Lorentz", index=False)
        lorentz_fit_data.to_excel(writer, sheet_name="Fitted Data", index=False)

    print("拟合完成。")


def modify_excel_names_to_concentration(file_path):
    """将文件名转换为浓度 (Index * 10)"""
    print("正在将文件名转换为浓度 (Index * 10)...")
    wb = load_workbook(file_path)

    def get_concentration_name(original_name):
        if not isinstance(original_name, str):
            return original_name
        match = re.search(r'(\d+)', original_name)
        if match:
            index = int(match.group(1))
            concentration = index * 10
            return str(concentration)
        return original_name

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name in ['All', 'Fitted Data']:
            for cell in sheet[1]:
                if cell.value and isinstance(cell.value, str):
                    if 'Wavelength' in cell.value or cell.value == 'x':
                        continue
                    new_val = get_concentration_name(cell.value)
                    if new_val != cell.value:
                        cell.value = new_val
                        cell.number_format = numbers.FORMAT_TEXT

        if sheet_name in ['Max_Value', 'Lorentz']:
            for row in sheet.iter_rows(min_row=2):
                cell = row[0]
                if cell.value:
                    new_val = get_concentration_name(cell.value)
                    if new_val != cell.value:
                        cell.value = new_val
                        cell.number_format = numbers.FORMAT_TEXT

    wb.save(file_path)
    print("浓度重命名完成。")


def main():
    print("欢迎使用荧光数据分析程序 (660.txt Version)！")
    print("注意：仅处理 660.txt 文件")

    folder_path = select_folder()
    output_filename = get_output_filename()
    output_file = os.path.join(folder_path, output_filename)

    try:
        aggregate_fluorescence_data(folder_path, output_file)
        extract_max_values(output_file)
        calculate_intensity_ratios(output_file)
        process_and_fit_data(output_file)
        modify_excel_names_to_concentration(output_file)

        print("正在格式化Excel...")
        wb = load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            format_fonts(wb[sheet_name])
        wb.save(output_file)

        print(f"\n【处理成功】结果已保存到：{output_file}")

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"\n处理出错：{str(e)}")

    print("\n按回车键退出...")
    input()


if __name__ == "__main__":
    main()