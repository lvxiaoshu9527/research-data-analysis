import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import tkinter as tk
from tkinter import filedialog
import sys
from scipy.optimize import curve_fit
import shutil
from collections import defaultdict

# --- 1. 全局设置与映射表 ---

# 【您的配置区】
# 1.1 定义总任务清单所需的所有手性 (标准格式)
ALL_CHIRALITIES = sorted(['(6,5)', '(8,3)', '(7,5)', '(6,5)-S7'])

# 1.2 定义18种氨基酸的标准三字母代码
ALL_AMINO_ACIDS = sorted([
    'Ala', 'Arg', 'Asn', 'Asp', 'Cys', 'Gln', 'Glu', 'Gly', 'His',
    'Ile', 'Leu', 'Lys', 'Met', 'Phe', 'Pro', 'Ser', 'Trp', 'Tyr'
])

# 1.3 定义别名/拼音到标准氨基酸名称的映射
PINYIN_MAP = {
    'ben': 'Phe', 'benbin': 'Phe',  # 苯丙氨酸
    'ban': 'Cys',  # 半胱氨酸
    'gan': 'Gly',  # 甘氨酸
    # 在这里继续添加...
}

# 1.4 定义文件夹名到标准手性格式的映射
CHIRAL_MAP = {
    "65-s7": "(6,5)-S7", "s7-65": "(6,5)-S7",
    "65": "(6,5)", "ss65": "(6,5)",
    "83": "(8,3)", "ss83": "(8,3)",
    "75": "(7,5)", "ss75": "(7,5)",
    "64": "(6,4)", "ss64": "(6,4)",
    "73": "(7,3)",
    "91": "(9,1)"
}


# --- 2. 核心处理函数 (拟合、格式化等) ---
# 这部分函数与之前版本相同，保持不变

def lorentzian(x, amp, cen, wid):
    return (amp * wid ** 2 / ((x - cen) ** 2 + wid ** 2))


def aggregate_fluorescence_data(directory_path, temp_output_file):
    print("  步骤 1/6: 正在汇总荧光数据...")
    data_frames = []
    for filename in sorted(os.listdir(directory_path)):
        if filename.endswith("660.txt"):
            file_path = os.path.join(directory_path, filename)
            df = pd.read_csv(file_path, sep='\t', header=None, names=["Wavelength", "Intensity"])
            df = df[(df['Wavelength'] >= 900) & (df['Wavelength'] <= 1400)]
            df = df.set_index('Wavelength').rename(columns={"Intensity": os.path.splitext(filename)[0]})
            data_frames.append(df)
    if not data_frames: return False
    all_data = pd.concat(data_frames, axis=1)
    all_data.index.name = "Wavelength"
    with pd.ExcelWriter(temp_output_file, engine='openpyxl') as writer:
        all_data.to_excel(writer, sheet_name='All')
    return True


def extract_max_values(file_path):
    print("  步骤 2/6: 正在提取峰值数据...")
    df = pd.read_excel(file_path, sheet_name='All', index_col='Wavelength')
    results = []
    for col in df.columns:
        if not df[col].empty:
            results.append({'Name': col, 'Peak Position': df[col].idxmax(), 'Intensity': df[col].max() - df[col].min()})
    new_df = pd.DataFrame(results)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        new_df.to_excel(writer, sheet_name='Max_Value', index=False)


def calculate_intensity_ratios(file_path):
    print("  步骤 3/6: 正在计算强度比值...")
    wb = load_workbook(file_path)
    if 'Max_Value' not in wb.sheetnames: return
    sheet = wb['Max_Value']
    data_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            try:
                data_list.append({'Name': str(row[0]), 'Intensity': float(row[2])})
            except (ValueError, TypeError):
                continue
    if not data_list: return
    df = pd.DataFrame(data_list)
    ref_samples = df[df['Name'].str.startswith("0__660", na=False)]
    if ref_samples.empty: return
    i0 = ref_samples['Intensity'].mean()
    if i0 == 0: return
    df['(I-I0)/I0'] = (df['Intensity'] - i0) / i0
    sheet['D1'] = '(I-I0)/I0'
    for i, ratio in enumerate(df['(I-I0)/I0'], start=2): sheet[f'D{i}'] = ratio
    wb.save(file_path)


def process_sheet_data(df, wave_range):
    mask = (df.index >= wave_range[0]) & (df.index <= wave_range[1])
    filtered_df = df.loc[mask]
    x = filtered_df.index.values
    results = pd.DataFrame()
    fit_data = pd.DataFrame({'Wavelength': x})
    for column in filtered_df.columns:
        y = filtered_df[column].values
        if len(y) < 3: continue
        try:
            max_y_index = np.argmax(y)
            p0 = [y[max_y_index], x[max_y_index], (x[-1] - x[0]) / 10]
            popt, _ = curve_fit(lorentzian, x, y, p0=p0, maxfev=8000)
            new_row = pd.DataFrame([{"Column": column, "Amplitude": popt[0], "Center": popt[1], "Width": popt[2]}])
            results = pd.concat([results, new_row], ignore_index=True)
            fit_data[column] = lorentzian(x, *popt)
        except Exception:
            continue
    return results, fit_data


def process_and_fit_data(file_path, chiral_key):
    print("  步骤 4/6: 正在进行洛伦兹拟合...")
    ranges = {"(6,4)": (871.2951, 961.1901), "(6,5)-S7": (895, 1100.975), "(6,5)": (950, 1100),
              "(7,3)": (961.1901, 1074), "(7,5)": (980.129, 1150.973), "(8,3)": (926.6856, 1050.47),
              "(9,1)": (890.3606, 1021.353)}
    wave_range = ranges.get(chiral_key)
    if not wave_range:
        print(f"    [警告] 未找到手性 '{chiral_key}' 的波长范围，跳过拟合。")
        return
    df = pd.read_excel(file_path, sheet_name='All', index_col='Wavelength')
    results, fit_data = process_sheet_data(df, wave_range)
    if not results.empty:
        ref_center = results['Center'].iloc[0]
        results['Δλ'] = results['Center'] - ref_center
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        results.to_excel(writer, sheet_name="Lorentz", index=False)
        fit_data.to_excel(writer, sheet_name="Fitted_Data", index=False)


def modify_excel_data(file_path, data_type):
    print("  步骤 5/6: 正在根据浓度类型修改数据...")
    wb = load_workbook(file_path)
    replacements = ({'0__660': '0', '1__660': '10', '2__660': '30', '3__660': '50', '4__660': '70', '5__660': '90',
                     '6__660': '110'} if data_type == 'predict' else {'0__660': '0', '1__660': '20', '2__660': '40',
                                                                      '3__660': '60', '4__660': '80', '5__660': '100',
                                                                      '6__660': '120'})
    sheets_to_process = {'All': 'row', 'Max_Value': 'column', 'Lorentz': 'column', 'Fitted_Data': 'row'}
    for sheet_name, direction in sheets_to_process.items():
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            iterator = sheet[1] if direction == 'row' else (row[0] for row in sheet.iter_rows(min_row=2))
            for cell in iterator:
                if isinstance(cell.value, str) and cell.value in replacements:
                    cell.value = replacements[cell.value]
    wb.save(file_path)


def format_excel_file(file_path):
    print("  步骤 6/6: 正在格式化 Excel...")
    wb = load_workbook(file_path)
    header_font, body_font = Font(name='Arial', bold=True, size=11), Font(name='Times New Roman', size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row == 1 else body_font
                cell.alignment = center_alignment
                cell.border = no_border
    wb.save(file_path)


# --- 3. 全新的元数据识别与流程控制函数 ---

def select_folder(title="选择文件夹"):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title=title)
    if not folder_path:
        print("没有选择文件夹，程序退出。")
        sys.exit()
    return folder_path


def get_completed_tasks(output_root_folder):
    """扫描输出文件夹，返回一个包含已完成任务元组(chiral, type, aa)的集合。"""
    completed_tasks = set()
    if not os.path.isdir(output_root_folder):
        return completed_tasks

    dir_pattern = re.compile(r"(\(.*\))-(train|predict)")
    file_pattern = re.compile(r"(\(.*\))output_data-([A-Za-z]{3})\d*\.xlsx")

    for dirpath, _, filenames in os.walk(output_root_folder):
        subfolder_name = os.path.basename(dirpath)
        dir_match = dir_pattern.match(subfolder_name)
        if not dir_match: continue

        chiral, data_type = dir_match.groups()
        for fname in filenames:
            file_match = file_pattern.match(fname)
            if file_match and file_match.group(1) == chiral:
                amino_acid = file_match.group(2)
                completed_tasks.add((chiral, data_type, amino_acid))
    return completed_tasks


def find_source_data_folders(root_dir):
    """递归查找所有包含'*660.txt'文件的文件夹。"""
    target_folders = []
    for dirpath, _, filenames in os.walk(root_dir):
        if any(fname.endswith("660.txt") for fname in filenames):
            target_folders.append(dirpath)
    return sorted(target_folders)


def get_folder_metadata(folder_path):
    """【升级版】分析文件夹完整路径，提取手性、氨基酸和类型信息。"""
    path_parts = folder_path.lower().split(os.sep)

    found_chiral, found_aa = None, None
    found_type = 'train'  # 默认是train类型

    for part in reversed(path_parts):
        # 识别类型
        if 'conc' in part:
            found_type = 'predict'

        # 识别手性
        if not found_chiral:
            cleaned_part = re.sub(r'ss|l-', '', part)  # 清理常见前缀
            if cleaned_part in CHIRAL_MAP:
                found_chiral = CHIRAL_MAP[cleaned_part]

        # 识别氨基酸
        if not found_aa:
            cleaned_part = re.sub(r'l-', '', part)
            if cleaned_part in PINYIN_MAP:
                found_aa = PINYIN_MAP[cleaned_part]
            elif cleaned_part.capitalize() in ALL_AMINO_ACIDS:
                found_aa = cleaned_part.capitalize()

    if found_chiral and found_aa:
        return {'chiral_key': found_chiral, 'amino_acid': found_aa, 'type': found_type}
    return None


def final_report(still_missing_tasks):
    """打印最终的缺失任务报告。"""
    print("\n\n--- 最终任务完成情况报告 ---")
    if not still_missing_tasks:
        print("🎉 恭喜！所有预定任务均已完成！")
        return

    print(f"⚠️ 注意：有 {len(still_missing_tasks)} 个预定任务因未在源数据中找到对应文件而未能完成。")

    # 按手性和类型分组报告
    report = defaultdict(list)
    for chiral, data_type, aa in sorted(list(still_missing_tasks)):
        report[(chiral, data_type)].append(aa)

    for (chiral, data_type), aas in report.items():
        print(f"\n【{chiral} - {data_type.upper()}组】缺失以下氨基酸:")
        print(f"  -> {', '.join(aas)}")


# --- 4. 主函数 ---
def main():
    print("欢迎使用荧光数据批量分析程序（V4 - 任务驱动终版）！")

    root_folder = select_folder(title="第一步：请选择包含【原始数据】的根目录")
    output_root_folder = select_folder(title="第二步：请选择用于保存【所有结果】的输出目录")

    # 1. 定义总任务清单
    required_tasks = set()
    for chiral in ALL_CHIRALITIES:
        for data_type in ['train', 'predict']:
            for aa in ALL_AMINO_ACIDS:
                required_tasks.add((chiral, data_type, aa))
    print(f"\n--- 阶段 1: 任务定义 ---")
    print(f"根据您的配置，总共需要完成 {len(required_tasks)} 个数据处理任务。")

    # 2. 评估现状，确定待办清单
    completed_tasks = get_completed_tasks(output_root_folder)
    tasks_to_do = required_tasks - completed_tasks
    print(f"\n--- 阶段 2: 现状评估 ---")
    print(f"在输出目录中找到 {len(completed_tasks)} 个已完成任务。")
    print(f"--> 还需处理 {len(tasks_to_do)} 个新任务。")

    if not tasks_to_do:
        print("\n所有任务均已完成，无需执行操作。")
        input("按回车键退出程序...")
        return

    # 3. 扫描源数据，开始处理
    print("\n--- 阶段 3: 扫描并处理待办任务 ---")
    all_source_folders = find_source_data_folders(root_folder)
    print(f"在源目录中找到 {len(all_source_folders)} 个含光谱文件的文件夹，开始匹配...")

    processed_count = 0
    naming_counter = defaultdict(lambda: defaultdict(int))  # 用于文件名编号

    for folder_path in all_source_folders:
        metadata = get_folder_metadata(folder_path)

        if not metadata:
            continue  # 无法识别路径的直接跳过

        task_tuple = (metadata['chiral_key'], metadata['type'], metadata['amino_acid'])

        if task_tuple in tasks_to_do:
            print(f"\n-> 发现待办任务: {task_tuple[0]}, {task_tuple[2]}, {task_tuple[1]}")
            temp_excel_file = os.path.join(output_root_folder, "__temp_analysis.xlsx")

            try:
                if not aggregate_fluorescence_data(folder_path, temp_excel_file): continue
                extract_max_values(temp_excel_file)
                calculate_intensity_ratios(temp_excel_file)
                process_and_fit_data(temp_excel_file, task_tuple[0])
                modify_excel_data(temp_excel_file, task_tuple[1])
                format_excel_file(temp_excel_file)

                chiral, data_type, aa = task_tuple
                count = naming_counter[chiral][data_type] + 1
                naming_counter[chiral][data_type] = count

                output_subfolder = os.path.join(output_root_folder, f"{chiral}-{data_type}")
                os.makedirs(output_subfolder, exist_ok=True)
                final_filename = f"{chiral}output_data-{aa}{count}.xlsx"
                final_filepath = os.path.join(output_subfolder, final_filename)

                shutil.move(temp_excel_file, final_filepath)
                print(f"  处理完成！结果已保存至: {final_filepath}")

                processed_count += 1
                tasks_to_do.remove(task_tuple)  # 从待办清单中移除，防止重复处理

            except Exception as e:
                print(f"  --- !!! 处理此文件夹时发生严重错误: {e} !!! ---")
                import traceback
                traceback.print_exc()
            finally:
                if os.path.exists(temp_excel_file):
                    os.remove(temp_excel_file)

    final_report(tasks_to_do)
    print(f"\n本次运行共处理并新增了 {processed_count} 个数据文件。")
    input("按回车键退出程序...")


if __name__ == "__main__":
    main()
